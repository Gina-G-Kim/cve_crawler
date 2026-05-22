"""
Space security CVE crawler.

This version follows the example schemas in example/cwe_software.csv and
example/emb3d-stix-2.0.1.json, while enriching each CVE from public sources
so the exported records carry as much real data as possible.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import logging
import re
import time
import uuid
from collections import Counter, defaultdict, deque
from dataclasses import dataclass, field
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Deque, Dict, Iterable, List, Optional, Sequence, Set, Tuple
from urllib.parse import parse_qsl, quote_plus, urlencode, urljoin, urlparse, urlunparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


LOGGER = logging.getLogger("space_cve_crawler")

SOURCE_URLS = {
    "github_poc": "https://github.com/ericyoc/prob_vuln_assess_space_iot_sys_poc",
    "opencve_satellite": "https://app.opencve.io/cve/?q=product%3Asatellite",
    "vulners_redhat": "https://vulners.com/search/vendors/redhat/products/satellite/versions",
    "enisa_space": "https://github.com/enisaeu/Space-Threat-Landscape",
}

CSV_HEADERS = [
    "type",
    "id",
    "created",
    "modified",
    "name",
    "description",
    "x_cve_id",
    "x_cve_severity",
    "x_cve_cvss_score",
    "x_cve_cvss_vector",
    "x_cve_cwe_ids",
    "x_cve_vendor",
    "x_cve_product",
    "x_cve_affected_versions",
    "x_cve_references",
    "x_cve_source_urls",
    "x_cve_source_names",
    "x_cve_mitigation",
    "x_cve_impact",
    "x_cve_exploitability",
    "x_cve_epss_score",
    "x_cve_epss_percentile",
    "x_cve_space_related",
    "x_cve_notes",
    "external_references",
]

def tree_line(label: str, description: str = "", indent: int = 0) -> str:
    prefix = "  " * indent
    if description:
        return f"{prefix}- {label} // {description}"
    return f"{prefix}- {label}"


def build_report_tree_lines() -> List[str]:
    lines = ["CSV / JSON 필드 트리"]
    csv_fields = [
        ("type", "레코드 유형"),
        ("id", "vulnerability 식별자"),
        ("created", "생성 시각"),
        ("modified", "수정 시각"),
        ("name", "취약점 이름"),
        ("description", "취약점 설명"),
        ("x_cve_id", "CVE 식별자"),
        ("x_cve_severity", "심각도"),
        ("x_cve_cvss_score", "CVSS 점수"),
        ("x_cve_cvss_vector", "CVSS 벡터"),
        ("x_cve_cwe_ids", "CWE ID 목록"),
        ("x_cve_vendor", "공급업체"),
        ("x_cve_product", "제품"),
        ("x_cve_affected_versions", "영향받는 버전 목록"),
        ("x_cve_references", "참고 링크 목록"),
        ("x_cve_source_urls", "수집 소스 URL 목록"),
        ("x_cve_source_names", "수집 소스 이름 목록"),
        ("x_cve_mitigation", "대응 방안"),
        ("x_cve_impact", "영향"),
        ("x_cve_exploitability", "공격 가능성"),
        ("x_cve_epss_score", "EPSS 점수"),
        ("x_cve_epss_percentile", "EPSS 백분위"),
        ("x_cve_space_related", "우주보안 관련 여부"),
        ("x_cve_notes", "메모"),
        ("external_references", "외부 참고 목록"),
    ]
    for name, description in csv_fields:
        lines.append(tree_line(name, description, 1))

    lines.append(tree_line("JSON bundle", "출력 번들 최상위 구조", 0))
    lines.append(tree_line("type", "번들 유형", 1))
    lines.append(tree_line("id", "번들 식별자", 1))
    lines.append(tree_line("x_cve_source_sections", "소스별 섹션 목록", 1))
    lines.append(tree_line("source_name", "소스 이름", 2))
    lines.append(tree_line("record_count", "레코드 수", 2))
    lines.append(tree_line("objects", "객체 목록", 1))

    lines.append(tree_line("identity 객체", "최상위 식별 객체", 2))
    for name, description in [
        ("type", "레코드 유형"),
        ("id", "identity 식별자"),
        ("created", "생성 시각"),
        ("modified", "수정 시각"),
        ("name", "이름"),
        ("identity_class", "식별자 분류"),
        ("x_cve_export_format", "내보내기 형식"),
    ]:
        lines.append(tree_line(name, description, 3))

    lines.append(tree_line("x-cve-section 객체", "소스별 묶음 객체", 2))
    for name, description in [
        ("type", "레코드 유형"),
        ("id", "x-cve-section 식별자"),
        ("name", "소스 이름"),
        ("description", "섹션 설명"),
        ("record_count", "레코드 수"),
    ]:
        lines.append(tree_line(name, description, 3))

    lines.append(tree_line("vulnerability 객체", "취약점 본문 객체", 2))
    vuln_fields = [
        ("type", "레코드 유형"),
        ("id", "vulnerability 식별자"),
        ("created", "생성 시각"),
        ("modified", "수정 시각"),
        ("name", "취약점 이름"),
        ("description", "취약점 설명"),
        ("x_cve_id", "CVE 식별자"),
        ("x_cve_severity", "심각도"),
        ("x_cve_cvss_score", "CVSS 점수"),
        ("x_cve_cvss_vector", "CVSS 벡터"),
        ("x_cve_cwe_ids", "CWE ID 목록"),
        ("x_cve_vendor", "공급업체"),
        ("x_cve_product", "제품"),
        ("x_cve_affected_versions", "영향받는 버전 목록"),
        ("x_cve_references", "참고 링크 목록"),
        ("x_cve_source_urls", "수집 소스 URL 목록"),
        ("x_cve_source_names", "수집 소스 이름 목록"),
        ("x_cve_mitigation", "대응 방안"),
        ("x_cve_impact", "영향"),
        ("x_cve_exploitability", "공격 가능성"),
        ("x_cve_epss_score", "EPSS 점수"),
        ("x_cve_epss_percentile", "EPSS 백분위"),
        ("x_cve_space_related", "우주보안 관련 여부"),
        ("x_cve_notes", "메모"),
        ("external_references", "외부 참고 목록"),
    ]
    for name, description in vuln_fields:
        lines.append(tree_line(name, description, 3))
    return lines


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def unique_keep_order(values: Iterable[str]) -> List[str]:
    seen: Set[str] = set()
    result: List[str] = []
    for value in values:
        value = clean_text(value)
        if not value or value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def join_values(values: Iterable[str], separator: str = "; ") -> str:
    return separator.join(unique_keep_order(values))


def normalize_url(url: str) -> str:
    url = clean_text(url)
    if url.startswith("//"):
        url = "https:" + url
    parsed = urlparse(url)
    query = urlencode(sorted(parse_qsl(parsed.query, keep_blank_values=True)))
    cleaned = parsed._replace(fragment="", query=query)
    return urlunparse(cleaned).rstrip("/")


def same_host(url: str, other: str) -> bool:
    return urlparse(url).netloc.lower() == urlparse(other).netloc.lower()


def is_http_url(url: str) -> bool:
    parsed = urlparse(url)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def github_blob_to_raw_url(url: str) -> str:
    parsed = urlparse(normalize_url(url))
    if parsed.netloc.lower() != "github.com":
        return normalize_url(url)
    parts = parsed.path.strip("/").split("/")
    if len(parts) >= 5 and parts[2] == "blob":
        owner, repo, _, branch = parts[:4]
        file_path = "/".join(parts[4:])
        return f"https://raw.githubusercontent.com/{owner}/{repo}/{branch}/{file_path}"
    return normalize_url(url)


def cve_ids_from_text(text: str) -> Set[str]:
    return set(re.findall(r"CVE-\d{4}-\d{4,7}", text or ""))


def snippet_around(text: str, needle: str, radius: int = 160) -> str:
    index = text.find(needle)
    if index < 0:
        return clean_text(text[: radius * 2])
    start = max(0, index - radius)
    end = min(len(text), index + len(needle) + radius)
    return clean_text(text[start:end])


def severity_label(score: Optional[float]) -> str:
    if score is None:
        return ""
    if score >= 9.0:
        return "Critical"
    if score >= 7.0:
        return "High"
    if score >= 4.0:
        return "Medium"
    if score > 0:
        return "Low"
    return "None"


def epss_label(score: Optional[float]) -> str:
    if score is None:
        return ""
    if score >= 0.7:
        return "High"
    if score >= 0.3:
        return "Medium"
    if score > 0:
        return "Low"
    return ""


def parse_cpe_uri(cpe: str) -> Tuple[str, str, str]:
    parts = cpe.split(":")
    if len(parts) < 6:
        return "", "", ""
    return parts[3], parts[4], parts[5]


def format_cpe_summary(cpes: Iterable[str]) -> str:
    entries = []
    for cpe in unique_keep_order(cpes):
        vendor, product, version = parse_cpe_uri(cpe)
        if vendor and product:
            item = f"{vendor}/{product}"
            if version and version not in {"*", "-"}:
                item += f" {version}"
            entries.append(item)
    return join_values(entries)


def derive_attack_surface(cvss_vector: str) -> str:
    if not cvss_vector:
        return ""
    parts = {chunk.split(":", 1)[0]: chunk.split(":", 1)[1] for chunk in cvss_vector.split("/") if ":" in chunk}
    mapping = {
        "AV": {"N": "Network", "A": "Adjacent", "L": "Local", "P": "Physical"},
        "AC": {"L": "Low complexity", "H": "High complexity"},
        "PR": {"N": "No privileges", "L": "Low privileges", "H": "High privileges"},
        "UI": {"N": "No user interaction", "R": "User interaction"},
    }
    fragments = []
    for key, options in mapping.items():
        value = parts.get(key)
        if value:
            fragments.append(options.get(value, value))
    return "; ".join(fragments)


def derive_impact(cvss: Dict[str, Any]) -> str:
    impacts = []
    for metric, label in [("confidentialityImpact", "Confidentiality"), ("integrityImpact", "Integrity"), ("availabilityImpact", "Availability")]:
        value = cvss.get(metric)
        if value and value != "NONE":
            impacts.append(f"{label}: {value}")
    return "; ".join(impacts)


def derive_cwe_ids(nvd_cwes: Iterable[Dict[str, Any]]) -> List[str]:
    values = []
    for item in nvd_cwes:
        if not isinstance(item, dict):
            continue
        for desc in item.get("description", []):
            value = desc.get("value")
            if value and value.startswith("CWE-"):
                values.append(value)
    return unique_keep_order(values)


def parse_cna_affect(affected: Iterable[Dict[str, Any]]) -> Tuple[List[str], List[str], List[str], List[str]]:
    vendors: List[str] = []
    products: List[str] = []
    versions: List[str] = []
    platforms: List[str] = []
    for item in affected:
        if not isinstance(item, dict):
            continue
        vendor = item.get("vendor")
        product = item.get("product")
        if vendor and vendor not in {"n/a", "*", "-"}:
            vendors.append(str(vendor))
        if product and product not in {"n/a", "*", "-"}:
            products.append(str(product))
        version_entries = item.get("versions", [])
        for version_entry in version_entries:
            if not isinstance(version_entry, dict):
                continue
            version = version_entry.get("version")
            if version and version not in {"n/a", "*", "-"}:
                versions.append(str(version))
            status = version_entry.get("status")
            if status:
                platforms.append(f"{product or 'product'}: {status}")
    return unique_keep_order(vendors), unique_keep_order(products), unique_keep_order(versions), unique_keep_order(platforms)


def collect_nvd_cpes(configurations: Any) -> List[str]:
    cpes: List[str] = []
    if isinstance(configurations, dict):
        nodes = configurations.get("nodes", [])
        for node in nodes:
            if not isinstance(node, dict):
                continue
            for match in node.get("cpeMatch", []):
                if isinstance(match, dict) and match.get("vulnerable") and match.get("criteria"):
                    cpes.append(match["criteria"])
    elif isinstance(configurations, list):
        for config in configurations:
            if not isinstance(config, dict):
                continue
            nodes = config.get("nodes", [])
            for node in nodes:
                if not isinstance(node, dict):
                    continue
                for match in node.get("cpeMatch", []):
                    if isinstance(match, dict) and match.get("vulnerable") and match.get("criteria"):
                        cpes.append(match["criteria"])
    return unique_keep_order(cpes)


def build_cache_key(value: str) -> str:
    return hashlib.sha1(value.encode("utf-8")).hexdigest()


class HttpClient:
    def __init__(self, timeout: int = 20, retries: int = 3, cache_dir: str = ".cache/http"):
        self.timeout = timeout
        self.retries = retries
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124 Safari/537.36",
            "Accept": "text/html,application/json;q=0.9,*/*;q=0.8",
        })
        self.cache_dir = Path(cache_dir)
        self.cache_dir.mkdir(parents=True, exist_ok=True)

    def _cache_path(self, url: str, suffix: str) -> Path:
        return self.cache_dir / f"{build_cache_key(url)}{suffix}"

    def get_text(self, url: str, use_cache: bool = True) -> str:
        url = normalize_url(url)
        cache_path = self._cache_path(url, ".txt")
        if use_cache and cache_path.exists():
            LOGGER.debug("HTTP cache hit (text): %s", url)
            return cache_path.read_text(encoding="utf-8")

        last_error: Optional[Exception] = None
        for attempt in range(self.retries):
            try:
                LOGGER.debug("HTTP GET (text) attempt %d/%d: %s", attempt + 1, self.retries, url)
                response = self.session.get(url, timeout=self.timeout)
                response.raise_for_status()
                text = response.text
                if use_cache:
                    cache_path.write_text(text, encoding="utf-8")
                LOGGER.debug("HTTP GET (text) success: %s (%d bytes)", url, len(text.encode("utf-8")))
                return text
            except Exception as exc:
                last_error = exc
                LOGGER.debug("HTTP GET (text) failed on attempt %d/%d for %s: %s", attempt + 1, self.retries, url, exc)
                time.sleep(1.5 ** attempt)
        raise RuntimeError(f"Failed to fetch {url}: {last_error}")

    def get_bytes(self, url: str, use_cache: bool = True) -> bytes:
        url = normalize_url(url)
        cache_path = self._cache_path(url, ".bin")
        if use_cache and cache_path.exists():
            LOGGER.debug("HTTP cache hit (bytes): %s", url)
            return cache_path.read_bytes()

        last_error: Optional[Exception] = None
        for attempt in range(self.retries):
            try:
                LOGGER.debug("HTTP GET (bytes) attempt %d/%d: %s", attempt + 1, self.retries, url)
                response = self.session.get(url, timeout=self.timeout)
                response.raise_for_status()
                content = response.content
                if use_cache:
                    cache_path.write_bytes(content)
                LOGGER.debug("HTTP GET (bytes) success: %s (%d bytes)", url, len(content))
                return content
            except Exception as exc:
                last_error = exc
                LOGGER.debug("HTTP GET (bytes) failed on attempt %d/%d for %s: %s", attempt + 1, self.retries, url, exc)
                time.sleep(1.5 ** attempt)
        raise RuntimeError(f"Failed to fetch {url}: {last_error}")

    def get_json(self, url: str, use_cache: bool = True) -> Dict[str, Any]:
        url = normalize_url(url)
        cache_path = self._cache_path(url, ".json")
        if use_cache and cache_path.exists():
            LOGGER.debug("HTTP cache hit (json): %s", url)
            return json.loads(cache_path.read_text(encoding="utf-8"))

        last_error: Optional[Exception] = None
        for attempt in range(self.retries):
            try:
                LOGGER.debug("HTTP GET (json) attempt %d/%d: %s", attempt + 1, self.retries, url)
                response = self.session.get(url, timeout=self.timeout)
                response.raise_for_status()
                payload = response.json()
                if use_cache:
                    cache_path.write_text(json.dumps(payload), encoding="utf-8")
                LOGGER.debug("HTTP GET (json) success: %s", url)
                return payload
            except Exception as exc:
                last_error = exc
                LOGGER.debug("HTTP GET (json) failed on attempt %d/%d for %s: %s", attempt + 1, self.retries, url, exc)
                time.sleep(1.5 ** attempt)
        raise RuntimeError(f"Failed to fetch {url}: {last_error}")


@dataclass
class Finding:
    cve_id: str
    source_name: str
    source_url: str
    page_url: str
    page_title: str
    snippet: str
    links: List[str] = field(default_factory=list)
    spec_version: Optional[str] = None


@dataclass
class CVERecord:
    cve_id: str
    spec_version: Optional[str] = None
    title: str = ""
    description: str = ""
    extended_description: str = ""
    severity: str = ""
    cvss_score: Optional[float] = None
    cvss_vector: str = ""
    epss_score: Optional[float] = None
    epss_percentile: Optional[float] = None
    cwe_ids: List[str] = field(default_factory=list)
    published_date: str = ""
    last_modified: str = ""
    vendor: str = ""
    product: str = ""
    affected_versions: List[str] = field(default_factory=list)
    affected_platforms: List[str] = field(default_factory=list)
    source_names: List[str] = field(default_factory=list)
    source_urls: List[str] = field(default_factory=list)
    references: List[str] = field(default_factory=list)
    evidence: List[str] = field(default_factory=list)
    mitigation: str = ""
    impact: str = ""
    exploitability: str = ""
    background_details: str = ""
    functional_area: str = ""
    alternate_terms: List[str] = field(default_factory=list)
    related_weaknesses: str = ""
    weakness_abstraction: str = "Base"
    status: str = "Draft"
    weakness_ordinalities: str = "Primary"
    modes_of_introduction: str = ""
    common_consequences: str = ""
    detection_methods: str = ""
    potential_mitigations: str = ""
    observed_examples: str = ""
    taxonomy_mappings: str = ""
    related_attack_patterns: str = ""
    notes: str = ""
    space_related: bool = True
    created_at: str = field(default_factory=utc_now_iso)

    def merge_finding(self, finding: Finding) -> None:
        self.source_names = unique_keep_order(self.source_names + [finding.source_name])
        self.source_urls = unique_keep_order(self.source_urls + [finding.source_url, finding.page_url])
        self.evidence = unique_keep_order(self.evidence + [finding.page_title, finding.snippet])

    def finalize(self) -> None:
        if not self.title:
            self.title = self.cve_id
        if not self.description:
            self.description = self.evidence[0] if self.evidence else self.title
        if not self.extended_description:
            self.extended_description = self.description
        if not self.product:
            if any("satellite" in source.lower() for source in self.source_names + [self.description, self.title]):
                self.product = "Satellite"
            else:
                self.product = "Space-related component"
        if not self.vendor:
            if any("red hat" in item.lower() for item in self.source_names + self.evidence):
                self.vendor = "Red Hat"
            elif any("github" in item.lower() for item in self.source_names):
                self.vendor = "GitHub"
            else:
                self.vendor = "Unknown"
        if not self.functional_area:
            self.functional_area = "Web UI; server-side component; supply chain ecosystem"
        if not self.background_details:
            self.background_details = "CVE collected from space-security-oriented source pages and enriched from public advisories."
        if not self.mitigation:
            self.mitigation = "Apply the vendor fix or upgrade to a patched version."
        if not self.potential_mitigations:
            self.potential_mitigations = self.mitigation
        if not self.exploitability:
            self.exploitability = derive_attack_surface(self.cvss_vector)
        if not self.common_consequences:
            self.common_consequences = self.impact
        if not self.observed_examples:
            self.observed_examples = join_values(self.references[:3])
        if not self.taxonomy_mappings:
            self.taxonomy_mappings = join_values(self.cwe_ids)
        if not self.alternate_terms:
            self.alternate_terms = ["satellite", "space", "space security"]
        if not self.notes:
            self.notes = join_values(self.evidence[:3], separator=" | ")

    @staticmethod
    def csv_headers() -> List[str]:
        return CSV_HEADERS

    def to_cwe_row(self) -> Dict[str, str]:
        return self.to_stix_object()

    @staticmethod
    def _csv_value(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        if isinstance(value, bool):
            return "true" if value else "false"
        if isinstance(value, (int, float)):
            return str(value)
        return str(value)

    def to_csv_row(self) -> Dict[str, str]:
        stix_object = self.to_stix_object()
        return {key: self._csv_value(stix_object.get(key)) for key in CSV_HEADERS}

    def to_stix_object(self) -> Dict[str, Any]:
        created = self.published_date or self.created_at
        modified = self.last_modified or self.created_at
        refs = [
            {
                "source_name": source,
                "url": url,
            }
            for source, url in zip(self.source_names or ["unknown"], self.source_urls or [""])
        ]
        obj: Dict[str, Any] = {
            "type": "vulnerability",
            "id": f"vulnerability--{uuid.uuid4()}",
            "created": created,
            "modified": modified,
            "name": self.title,
            "description": self.extended_description or self.description,
            "x_cve_id": self.cve_id,
            "x_cve_severity": self.severity,
            "x_cve_cvss_score": self.cvss_score,
            "x_cve_cvss_vector": self.cvss_vector,
            "x_cve_cwe_ids": self.cwe_ids,
            "x_cve_vendor": self.vendor,
            "x_cve_product": self.product,
            "x_cve_affected_versions": self.affected_versions,
            "x_cve_references": self.references,
            "x_cve_source_urls": self.source_urls,
            "x_cve_source_names": self.source_names,
            "x_cve_mitigation": self.mitigation,
            "x_cve_impact": self.impact,
            "x_cve_exploitability": self.exploitability,
            "x_cve_epss_score": self.epss_score,
            "x_cve_epss_percentile": self.epss_percentile,
            "x_cve_space_related": self.space_related,
            "x_cve_notes": self.notes,
            "external_references": refs,
        }
        # Only include spec_version if the source provided one
        if getattr(self, "spec_version", None):
            obj["spec_version"] = self.spec_version
        return obj


class BaseCrawler:
    def __init__(self, client: HttpClient, source_name: str, seed_url: str, max_depth: int = 2, max_pages: int = 12, log_fetch_errors: bool = True):
        self.client = client
        self.source_name = source_name
        self.seed_url = normalize_url(seed_url)
        self.max_depth = max_depth
        self.max_pages = max_pages
        self.log_fetch_errors = log_fetch_errors
        self.visited: Set[str] = set()

    def crawl(self) -> List[Finding]:
        queue: Deque[Tuple[str, int]] = deque([(self.seed_url, 0)])
        findings: List[Finding] = []
        LOGGER.info("%s crawl start: seed=%s max_depth=%d max_pages=%d", self.source_name, self.seed_url, self.max_depth, self.max_pages)

        while queue and len(self.visited) < self.max_pages:
            url, depth = queue.popleft()
            url = normalize_url(url)
            if url in self.visited or depth > self.max_depth:
                continue
            if not is_http_url(url):
                continue
            if not self.allow_url(url):
                continue

            self.visited.add(url)
            LOGGER.info("%s visiting depth=%d visited=%d url=%s", self.source_name, depth, len(self.visited), url)
            try:
                html = self.client.get_text(url)
            except Exception as exc:
                if self.log_fetch_errors:
                    LOGGER.warning("%s fetch failed for %s: %s", self.source_name, url, exc)
                continue

            # Attempt to detect JSON/STIX payloads and capture spec_version if present
            spec_ver: Optional[str] = None
            try:
                payload = json.loads(html)
                if isinstance(payload, dict):
                    # Top-level spec_version
                    spec_ver = payload.get("spec_version")
                    # Or search objects for a spec_version
                    if not spec_ver and isinstance(payload.get("objects"), list):
                        for obj in payload.get("objects", []):
                            if isinstance(obj, dict) and obj.get("spec_version"):
                                spec_ver = obj.get("spec_version")
                                break
            except Exception:
                spec_ver = None

            page_findings = self.extract_findings(html, url, spec_ver)
            findings.extend(page_findings)
            LOGGER.info("%s parsed %d findings from %s", self.source_name, len(page_findings), url)
            if depth < self.max_depth:
                discovered_links = self.extract_links(html, url)
                LOGGER.debug("%s discovered %d links from %s", self.source_name, len(discovered_links), url)
                for link in discovered_links:
                    if link not in self.visited and self.allow_url(link):
                        queue.append((link, depth + 1))

        LOGGER.info("%s crawl complete: visited=%d findings=%d", self.source_name, len(self.visited), len(findings))
        return findings

    def allow_url(self, url: str) -> bool:
        return same_host(url, self.seed_url)

    def extract_links(self, html: str, base_url: str) -> List[str]:
        soup = BeautifulSoup(html, "lxml")
        links = []
        for anchor in soup.find_all("a", href=True):
            href = normalize_url(urljoin(base_url, anchor["href"]))
            if is_http_url(href):
                links.append(href)
        return unique_keep_order(links)

    def extract_findings(self, html: str, page_url: str, spec_version: Optional[str] = None) -> List[Finding]:
        # HTML parsing still works even if the page is JSON text; pass through spec_version
        soup = BeautifulSoup(html, "lxml")
        page_title = clean_text(soup.title.get_text(" ", strip=True)) if soup.title else clean_text(urlparse(page_url).path.rsplit("/", 1)[-1])
        visible_text = clean_text(soup.get_text(" ", strip=True))
        findings = []
        for cve_id in cve_ids_from_text(visible_text):
            findings.append(
                Finding(
                    cve_id=cve_id,
                    source_name=self.source_name,
                    source_url=self.seed_url,
                    page_url=page_url,
                    page_title=page_title,
                    snippet=snippet_around(visible_text, cve_id),
                    spec_version=spec_version,
                )
            )
        return findings


class GitHubCrawler(BaseCrawler):
    def __init__(self, client: HttpClient, source_name: str, seed_url: str, max_depth: int = 2, max_pages: int = 12):
        super().__init__(client, source_name, seed_url, max_depth=max_depth, max_pages=max_pages)
        self.owner_repo = self._owner_repo(seed_url)

    @staticmethod
    def _owner_repo(url: str) -> str:
        path = urlparse(url).path.strip("/").split("/")
        return "/".join(path[:2]) if len(path) >= 2 else ""

    def allow_url(self, url: str) -> bool:
        parsed = urlparse(url)
        if parsed.netloc.lower() == "raw.githubusercontent.com":
            return self.owner_repo in parsed.path
        if parsed.netloc.lower() != "github.com":
            return False
        if not self.owner_repo:
            return False
        return parsed.path.startswith(f"/{self.owner_repo}")

    def crawl(self) -> List[Finding]:
        LOGGER.info("%s spreadsheet discovery starting for %s", self.source_name, self.seed_url)
        findings = self._crawl_xlsx_candidates()
        LOGGER.info("%s xlsx scan produced %d findings", self.source_name, len(findings))
        findings.extend(super().crawl())
        LOGGER.info("%s readme scan starting", self.source_name)
        findings.extend(self._crawl_readme_candidates())
        LOGGER.info("%s total findings after GitHub crawl: %d", self.source_name, len(findings))
        return findings

    def _discover_xlsx_links(self) -> List[str]:
        try:
            html = self.client.get_text(self.seed_url)
        except Exception as exc:
            LOGGER.warning("%s spreadsheet discovery failed for %s: %s", self.source_name, self.seed_url, exc)
            return []

        soup = BeautifulSoup(html, "lxml")
        links: List[str] = []
        for anchor in soup.find_all("a", href=True):
            href = anchor["href"]
            href_text = clean_text(anchor.get_text(" ", strip=True)).lower()
            if ".xlsx" not in href.lower() and ".xlsx" not in href_text:
                continue
            if href.startswith("/"):
                href = f"https://github.com{href}"
            href = normalize_url(urljoin(self.seed_url, href))
            if ".xlsx" in href.lower():
                links.append(href)
        unique_links = unique_keep_order(links)
        LOGGER.info("%s discovered %d xlsx link(s)", self.source_name, len(unique_links))
        LOGGER.debug("%s xlsx links: %s", self.source_name, unique_links)
        return unique_links

    def _crawl_xlsx_candidates(self) -> List[Finding]:
        findings: List[Finding] = []
        for link in self._discover_xlsx_links():
            raw_url = github_blob_to_raw_url(link)
            LOGGER.info("%s downloading xlsx: %s", self.source_name, raw_url)
            try:
                workbook_bytes = self.client.get_bytes(raw_url)
            except Exception as exc:
                LOGGER.warning("%s xlsx fetch failed for %s: %s", self.source_name, raw_url, exc)
                continue

            try:
                workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
            except Exception as exc:
                LOGGER.warning("%s xlsx parse failed for %s: %s", self.source_name, raw_url, exc)
                continue

            LOGGER.info("%s parsed workbook %s with %d sheet(s)", self.source_name, raw_url, len(workbook.worksheets))

            for sheet in workbook.worksheets:
                sheet_findings = 0
                row_count = 0
                for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    row_count += 1
                    row_text = clean_text(" | ".join(clean_text(cell) for cell in row if clean_text(cell)))
                    if not row_text:
                        continue
                    for cve_id in cve_ids_from_text(row_text):
                        sheet_findings += 1
                        findings.append(
                            Finding(
                                cve_id=cve_id,
                                source_name=self.source_name,
                                source_url=self.seed_url,
                                page_url=raw_url,
                                page_title=f"{self.owner_repo or self.source_name} / {sheet.title}",
                                snippet=snippet_around(row_text, cve_id),
                                spec_version=None,
                            )
                        )
                LOGGER.info("%s sheet %s rows=%d findings=%d", self.source_name, sheet.title, row_count, sheet_findings)
        return findings

    def _crawl_readme_candidates(self) -> List[Finding]:
        if not self.owner_repo:
            return []
        owner, repo = self.owner_repo.split("/", 1)
        findings: List[Finding] = []
        for branch in ("main", "master", "trunk"):
            for candidate in (
                f"https://raw.githubusercontent.com/{owner}/{repo}/{branch}/README.md",
                f"https://raw.githubusercontent.com/{owner}/{repo}/{branch}/readme.md",
            ):
                try:
                    text = self.client.get_text(candidate)
                except Exception:
                    continue
                LOGGER.info("%s readme candidate fetched: %s", self.source_name, candidate)
                for cve_id in cve_ids_from_text(text):
                    findings.append(
                        Finding(
                            cve_id=cve_id,
                            source_name=self.source_name,
                            source_url=self.seed_url,
                            page_url=candidate,
                            page_title=f"{repo} README",
                            snippet=snippet_around(text, cve_id),
                            spec_version=None,
                        )
                    )
                if findings:
                    break
        return findings

    def extract_links(self, html: str, base_url: str) -> List[str]:
        soup = BeautifulSoup(html, "lxml")
        links = []
        for anchor in soup.find_all("a", href=True):
            href = anchor["href"]
            if href.startswith("#"):
                continue
            if href.startswith("/"):
                href = f"https://github.com{href}"
            href = normalize_url(urljoin(base_url, href))
            parsed = urlparse(href)
            if parsed.netloc.lower() == "github.com" and self.owner_repo in parsed.path:
                if any(token in parsed.path for token in ("/issues/", "/blob/", "/tree/", "/releases", "/security", "/discussions")):
                    links.append(href)
            elif parsed.netloc.lower() == "raw.githubusercontent.com" and self.owner_repo in parsed.path:
                links.append(href)
        return unique_keep_order(links)


class OpenCVECrawler(BaseCrawler):
    def allow_url(self, url: str) -> bool:
        parsed = urlparse(url)
        return parsed.netloc.lower() == "app.opencve.io" and (parsed.path.startswith("/cve") or parsed.path.startswith("/"))

    def extract_links(self, html: str, base_url: str) -> List[str]:
        soup = BeautifulSoup(html, "lxml")
        links = []
        for anchor in soup.find_all("a", href=True):
            href = normalize_url(urljoin(base_url, anchor["href"]))
            parsed = urlparse(href)
            if parsed.netloc.lower() != "app.opencve.io":
                continue
            if parsed.path.startswith("/cve") or "page=" in parsed.query:
                links.append(href)
        return unique_keep_order(links)


class VulnersCrawler(BaseCrawler):
    def __init__(self, client: HttpClient, source_name: str, seed_url: str, max_depth: int = 2, max_pages: int = 12):
        super().__init__(client, source_name, seed_url, max_depth=max_depth, max_pages=max_pages, log_fetch_errors=False)

    def allow_url(self, url: str) -> bool:
        parsed = urlparse(url)
        return parsed.netloc.lower() == "vulners.com" and any(part in parsed.path for part in ("search", "vuln", "doc"))

    def crawl(self) -> List[Finding]:
        LOGGER.info("Vulners crawl start: %s", self.seed_url)
        findings = super().crawl()
        if findings:
            LOGGER.info("Vulners crawl succeeded without fallback: %d findings", len(findings))
            return findings
        fallback = self._crawl_nvd_fallback()
        if fallback:
            LOGGER.info("Vulners blocked; using NVD fallback for %s", self.seed_url)
        else:
            LOGGER.info("Vulners fallback returned no findings for %s", self.seed_url)
        return fallback

    def extract_links(self, html: str, base_url: str) -> List[str]:
        soup = BeautifulSoup(html, "lxml")
        links = []
        for anchor in soup.find_all("a", href=True):
            href = normalize_url(urljoin(base_url, anchor["href"]))
            parsed = urlparse(href)
            if parsed.netloc.lower() == "vulners.com" and any(part in parsed.path for part in ("search", "vuln", "doc")):
                links.append(href)
        return unique_keep_order(links)

    def _crawl_nvd_fallback(self) -> List[Finding]:
        query = quote_plus("red hat satellite")
        search_urls = [
            f"https://services.nvd.nist.gov/rest/json/cves/2.0?keywordSearch={query}",
            f"https://services.nvd.nist.gov/rest/json/cves/2.0?virtualMatchString={quote_plus('cpe:2.3:a:redhat:satellite:*:*:*:*:*:*:*:*')}",
        ]
        findings: List[Finding] = []
        seen: Set[str] = set()
        for search_url in search_urls:
            LOGGER.info("Vulners fallback querying NVD: %s", search_url)
            try:
                payload = self.client.get_json(search_url)
            except Exception as exc:
                LOGGER.warning("Vulners fallback search failed for %s: %s", search_url, exc)
                continue

            for vulnerability in payload.get("vulnerabilities", []):
                if not isinstance(vulnerability, dict):
                    continue
                cve = vulnerability.get("cve", {})
                if not isinstance(cve, dict):
                    continue
                cve_id = cve.get("id")
                if not cve_id or cve_id in seen:
                    continue
                seen.add(cve_id)
                descriptions = cve.get("descriptions", [])
                description = next((item.get("value", "") for item in descriptions if isinstance(item, dict) and item.get("lang") == "en"), "")
                references = [ref.get("url", "") for ref in cve.get("references", []) if isinstance(ref, dict) and ref.get("url")]
                findings.append(
                    Finding(
                        cve_id=cve_id,
                        source_name=f"{self.source_name} (fallback via NVD)",
                        source_url=search_url,
                        page_url=search_url,
                        page_title=clean_text(cve.get("id", cve_id)),
                        snippet=clean_text(description or join_values(references[:2])),
                        spec_version=None,
                    )
                )
            LOGGER.info("Vulners fallback query yielded %d unique findings so far", len(findings))
        return findings


class NVDEnricher:
    def __init__(self, client: HttpClient):
        self.client = client

    def enrich(self, cve_id: str) -> Dict[str, Any]:
        LOGGER.debug("Enrichment start: %s", cve_id)
        cna_url = f"https://cveawg.mitre.org/api/cve/{cve_id}"
        nvd_url = f"https://services.nvd.nist.gov/rest/json/cves/2.0?cveId={cve_id}"
        epss_url = f"https://api.first.org/data/v1/epss?cve={cve_id}"

        try:
            cna_payload = self.client.get_json(cna_url)
        except Exception:
            cna_payload = {}

        try:
            nvd_payload = self.client.get_json(nvd_url)
        except Exception:
            nvd_payload = {}

        try:
            epss_payload = self.client.get_json(epss_url)
        except Exception:
            epss_payload = {}

        cna = (cna_payload.get("containers") or {}).get("cna", {}) if isinstance(cna_payload, dict) else {}
        descriptions = cna.get("descriptions", []) if isinstance(cna, dict) else []
        description = next((item.get("value", "") for item in descriptions if isinstance(item, dict) and item.get("lang") == "en"), "")
        cna_refs = [ref.get("url", "") for ref in cna.get("references", []) if isinstance(ref, dict) and ref.get("url")]
        vendors, products, versions, platforms = parse_cna_affect(cna.get("affected", []))
        published_date = cna.get("datePublic", "") or cve_id
        last_modified = (cna_payload.get("cveMetadata", {}) or {}).get("dateUpdated", "") if isinstance(cna_payload, dict) else ""

        nvd_cve = {}
        if nvd_payload.get("vulnerabilities"):
            first_vulnerability = nvd_payload["vulnerabilities"][0]
            if isinstance(first_vulnerability, dict):
                nvd_cve = first_vulnerability.get("cve", {}) if isinstance(first_vulnerability.get("cve", {}), dict) else {}

        nvd_descriptions = nvd_cve.get("descriptions", []) if isinstance(nvd_cve, dict) else []
        nvd_description = next((item.get("value", "") for item in nvd_descriptions if isinstance(item, dict) and item.get("lang") == "en"), "")
        refs = unique_keep_order(cna_refs + [ref.get("url", "") for ref in nvd_cve.get("references", []) if isinstance(ref, dict) and ref.get("url")])
        weaknesses = derive_cwe_ids(nvd_cve.get("weaknesses", []))

        metrics = nvd_cve.get("metrics", {}) if isinstance(nvd_cve, dict) else {}
        metric_obj = {}
        for key in ("cvssMetricV31", "cvssMetricV30", "cvssMetricV2"):
            metric_list = metrics.get(key)
            if isinstance(metric_list, list) and metric_list:
                first_metric = metric_list[0]
                if isinstance(first_metric, dict):
                    metric_obj = first_metric.get("cvssData", {}) if isinstance(first_metric.get("cvssData", {}), dict) else {}
                    break

        score = metric_obj.get("baseScore")
        vector = metric_obj.get("vectorString", "")
        severity = metric_obj.get("baseSeverity") or severity_label(score)
        cvss_summary = metrics.get("cvssMetricV31") or metrics.get("cvssMetricV30") or metrics.get("cvssMetricV2") or [{}]
        cvss_data = cvss_summary[0].get("cvssData", {}) if isinstance(cvss_summary, list) and cvss_summary and isinstance(cvss_summary[0], dict) else {}

        cpes = collect_nvd_cpes(nvd_cve.get("configurations", [])) if isinstance(nvd_cve, dict) else []
        nvd_vendors = []
        nvd_products = []
        nvd_versions = []
        for cpe in cpes:
            vendor, product, version = parse_cpe_uri(cpe)
            if vendor and vendor not in {"*", "-"}:
                nvd_vendors.append(vendor)
            if product and product not in {"*", "-"}:
                nvd_products.append(product)
            if version and version not in {"*", "-"}:
                nvd_versions.append(version)

        if not vendors:
            vendors = unique_keep_order(nvd_vendors)
        if not products:
            products = unique_keep_order(nvd_products)
        if not versions:
            versions = unique_keep_order(nvd_versions)
        if not platforms and cpes:
            platforms = [format_cpe_summary(cpes)]

        epss = {}
        if isinstance(epss_payload, dict) and epss_payload.get("data"):
            data = epss_payload.get("data")
            if isinstance(data, list) and data and isinstance(data[0], dict):
                epss = data[0]

        primary_description = nvd_description or description

        LOGGER.debug(
            "Enrichment complete: %s severity=%s score=%s vendor=%s product=%s refs=%d",
            cve_id,
            severity,
            score,
            join_values(vendors),
            join_values(products),
            len(refs),
        )

        return {
            "title": cve_id,
            "description": primary_description,
            "extended_description": primary_description,
            "severity": severity,
            "cvss_score": score,
            "cvss_vector": vector,
            "cwe_ids": weaknesses,
            "published_date": published_date,
            "last_modified": last_modified,
            "vendor": join_values(vendors),
            "product": join_values(products),
            "affected_versions": unique_keep_order(versions),
            "affected_platforms": platforms,
            "references": refs,
            "impact": derive_impact(cvss_data),
            "exploitability": derive_attack_surface(vector),
            "epss_score": float(epss.get("epss")) if epss.get("epss") else None,
            "epss_percentile": float(epss.get("percentile")) if epss.get("percentile") else None,
            "modes_of_introduction": "Implementation",
            "status": "Stable" if primary_description and score is not None else "Draft",
            "detection_methods": "MITRE CVE JSON; NVD; public advisory review; source page review",
            "potential_mitigations": "Patch or upgrade affected components; apply vendor workarounds if available.",
            "common_consequences": derive_impact(cvss_data),
            "taxonomy_mappings": join_values(weaknesses or ["CVE"]),
            "related_attack_patterns": "",
            "related_weaknesses": join_values(weaknesses),
            "weakness_abstraction": "Base",
            "weakness_ordinalities": "Primary",
        }


class SpaceSecurityCrawler:
    def __init__(self, output_dir: str = "output", max_depth: int = 2, max_pages: int = 12):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.client = HttpClient()
        self.enricher = NVDEnricher(self.client)
        self.max_depth = max_depth
        self.max_pages = max_pages
        self.records: Dict[str, CVERecord] = {}
        self.findings_by_cve: Dict[str, List[Finding]] = defaultdict(list)
        self.sources_crawled: List[str] = []
        self.link_count = 0

    def crawl(self) -> List[CVERecord]:
        crawlers = [
            GitHubCrawler(self.client, "GitHub POC", SOURCE_URLS["github_poc"], self.max_depth, self.max_pages),
            OpenCVECrawler(self.client, "OpenCVE Satellite", SOURCE_URLS["opencve_satellite"], self.max_depth, self.max_pages),
            VulnersCrawler(self.client, "Vulners Red Hat Satellite", SOURCE_URLS["vulners_redhat"], self.max_depth, self.max_pages),
            GitHubCrawler(self.client, "ENISA Space Threat Landscape", SOURCE_URLS["enisa_space"], self.max_depth, self.max_pages),
        ]

        LOGGER.info("Crawler plan: %d sources", len(crawlers))
        for crawler in crawlers:
            LOGGER.info("Crawling %s from %s", crawler.source_name, crawler.seed_url)
            findings = crawler.crawl()
            self.sources_crawled.append(crawler.seed_url)
            self.link_count += len(crawler.visited)
            LOGGER.info("%s produced %d finding(s)", crawler.source_name, len(findings))
            for finding in findings:
                self.findings_by_cve[finding.cve_id].append(finding)

        for cve_id, findings in sorted(self.findings_by_cve.items()):
            LOGGER.info("Building record for %s from %d finding(s)", cve_id, len(findings))
            record = self.build_record(cve_id, findings)
            self.records[cve_id] = record

        LOGGER.info("Crawl aggregation complete: %d unique CVEs", len(self.records))
        return list(self.records.values())

    def build_record(self, cve_id: str, findings: List[Finding]) -> CVERecord:
        base = CVERecord(cve_id=cve_id)
        LOGGER.debug("Record build start: %s", cve_id)
        for finding in findings:
            base.merge_finding(finding)

        # If any finding provided a spec_version (from a JSON/STIX source), use it
        for f in findings:
            if getattr(f, "spec_version", None):
                base.spec_version = f.spec_version
                break

        enriched = {}
        try:
            enriched = self.enricher.enrich(cve_id)
        except Exception as exc:
            LOGGER.warning("Enrichment failed for %s: %s", cve_id, exc)

        for key, value in enriched.items():
            if hasattr(base, key):
                setattr(base, key, value)

        if not base.title and findings:
            base.title = findings[0].page_title or cve_id
        if not base.description and findings:
            base.description = findings[0].snippet
        if not base.extended_description:
            base.extended_description = base.description
        if not base.references:
            base.references = unique_keep_order([finding.page_url for finding in findings])
        if not base.source_urls:
            base.source_urls = unique_keep_order([finding.source_url for finding in findings])
        if not base.source_names:
            base.source_names = unique_keep_order([finding.source_name for finding in findings])
        if not base.evidence:
            base.evidence = unique_keep_order([finding.snippet for finding in findings])

        if not base.product:
            if any("satellite" in value.lower() for value in [base.title, base.description] + base.source_names):
                base.product = "Satellite"
            else:
                base.product = "Space-related component"

        if not base.vendor:
            if any("red hat" in value.lower() for value in [base.title, base.description] + base.evidence):
                base.vendor = "Red Hat"
            elif any("github" in value.lower() for value in base.source_names):
                base.vendor = "GitHub"
            elif any("enisa" in value.lower() for value in base.source_names):
                base.vendor = "ENISA"
            else:
                base.vendor = "Unknown"

        base.background_details = join_values([
            base.background_details,
            f"Sources: {join_values(base.source_names)}",
            f"Space-related context: {join_values(base.evidence[:2])}",
        ], separator=" | ")
        base.notes = join_values([
            base.notes,
            f"Source URLs: {join_values(base.source_urls)}",
        ], separator=" | ")
        base.space_related = True
        base.finalize()
        LOGGER.debug("Record build complete: %s sources=%s vendor=%s product=%s", cve_id, join_values(base.source_names), base.vendor, base.product)
        return base

    def export(self, records: Sequence[CVERecord]) -> Tuple[Path, Path, Path]:
        csv_path = self.output_dir / "space_cves.csv"
        json_path = self.output_dir / "space_cves.json"
        report_path = self.output_dir / "crawl_report.txt"

        self.write_csv(csv_path, records)
        self.write_json(json_path, records)
        self.write_report(report_path, records)
        return csv_path, json_path, report_path

    @staticmethod
    def _primary_source_name(record: CVERecord) -> str:
        if record.source_names:
            return record.source_names[0]
        return "Unknown source"

    def _source_counts(self, records: Sequence[CVERecord]) -> Dict[str, int]:
        counts: Dict[str, int] = {}
        for record in records:
            source_name = self._primary_source_name(record)
            counts[source_name] = counts.get(source_name, 0) + 1
        return counts

    def _build_json_summary(self, records: Sequence[CVERecord]) -> List[str]:
        counts = self._source_counts(records)
        total_records = len(records)
        min_cve = records[0].cve_id if records else "n/a"
        max_cve = records[-1].cve_id if records else "n/a"
        return [
            f"수집한 CVE/vulnerability: {total_records}",
            f"소스 구역: {len(counts)}",
            f"CVE 범위: {min_cve} ~ {max_cve}",
        ]

    def _build_json_objects(self, records: Sequence[CVERecord]) -> List[Dict[str, Any]]:
        grouped: Dict[str, List[CVERecord]] = defaultdict(list)
        for record in records:
            grouped[self._primary_source_name(record)].append(record)

        objects: List[Dict[str, Any]] = [
            {
                "type": "identity",
                "id": f"identity--{uuid.uuid4()}",
                "created": utc_now_iso(),
                "modified": utc_now_iso(),
                "name": "Space Security CVE Crawler",
                "identity_class": "organization",
                "x_cve_export_format": "cwe_software_style",
            }
        ]

        for source_name, grouped_records in grouped.items():
            objects.append(
                {
                    "type": "x-cve-section",
                    "id": f"x-cve-section--{uuid.uuid4()}",
                    "name": source_name,
                    "description": f"Records sourced primarily from {source_name}",
                    "record_count": len(grouped_records),
                }
            )
            objects.extend(record.to_stix_object() for record in grouped_records)

        return objects

    def write_csv(self, path: Path, records: Sequence[CVERecord]) -> None:
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=CSV_HEADERS)
            writer.writeheader()
            for record in records:
                writer.writerow(record.to_csv_row())

    def write_json(self, path: Path, records: Sequence[CVERecord]) -> None:
        bundle = {
            "type": "bundle",
            "id": f"bundle--{uuid.uuid4()}",
            "x_cve_source_sections": [
                {
                    "source_name": source_name,
                    "record_count": count,
                }
                for source_name, count in self._source_counts(records).items()
            ],
            "objects": self._build_json_objects(records),
        }

        path.write_text(json.dumps(bundle, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    def write_report(self, path: Path, records: Sequence[CVERecord]) -> None:
        severities = Counter(record.severity or "Unknown" for record in records)
        vendors = Counter(record.vendor or "Unknown" for record in records)
        with path.open("w", encoding="utf-8") as handle:
            handle.write("SPACE SECURITY CVE CRAWL REPORT\n")
            handle.write("=" * 72 + "\n\n")
            handle.write(f"Crawl time: {utc_now_iso()}\n")
            handle.write(f"Total CVEs: {len(records)}\n")
            handle.write(f"Sources crawled: {len(self.sources_crawled)}\n")
            handle.write(f"Links followed: {self.link_count}\n\n")
            handle.write("Sources:\n")
            for source in self.sources_crawled:
                handle.write(f"- {source}\n")
            handle.write("\nSeverity distribution:\n")
            for key, value in severities.most_common():
                handle.write(f"- {key}: {value}\n")
            handle.write("\nVendor distribution:\n")
            for key, value in vendors.most_common():
                handle.write(f"- {key}: {value}\n")
            handle.write("\nSample CVEs:\n")
            for record in list(records)[:10]:
                handle.write(f"- {record.cve_id}: {record.title}\n")
            handle.write("\nCSV / JSON 필드 트리:\n")
            for line in build_report_tree_lines():
                handle.write(f"{line}\n")


def configure_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
        handlers=[logging.StreamHandler(), logging.FileHandler("crawler.log", encoding="utf-8")],
    )


def run_crawler(output_dir: str = "output", max_depth: int = 2, max_pages: int = 12) -> List[CVERecord]:
    crawler = SpaceSecurityCrawler(output_dir=output_dir, max_depth=max_depth, max_pages=max_pages)
    records = crawler.crawl()
    records.sort(key=lambda record: record.cve_id)
    crawler.export(records)
    return records


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Space security CVE crawler")
    parser.add_argument("--output-dir", default="output", help="Directory for CSV/JSON output")
    parser.add_argument("--max-depth", type=int, default=2, help="Maximum link-following depth")
    parser.add_argument("--max-pages", type=int, default=12, help="Maximum pages per source")
    parser.add_argument("--log-level", default="INFO", choices=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"], help="Logging verbosity")
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = build_arg_parser().parse_args(argv)
    configure_logging()
    logging.getLogger().setLevel(getattr(logging, args.log_level))

    LOGGER.info("Starting space security CVE crawl at log level %s", args.log_level)
    records = run_crawler(output_dir=args.output_dir, max_depth=args.max_depth, max_pages=args.max_pages)
    LOGGER.info("Crawl complete: %d CVEs written to %s", len(records), args.output_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
